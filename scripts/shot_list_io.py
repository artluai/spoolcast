"""
shot_list_io.py — read/write the spoolcast shot list as xlsx.

The shot list has ONE sheet. Every column is visible without tab-switching.
Chunk-level columns use merged cells across the chunk's beats so grouping
is visually obvious. Rows are tinted per chunk (cycling pastel palette).

Layout (top-down):
  row 1: title
  row 2: session metadata (session id, canvas aspect, fps)
  row 3: separator (blank)
  row 4: table headers
  row 5+: beats (rows grouped by chunk via merged cells + fill color)

Usage (as a library):
    import shot_list_io
    data = shot_list_io.load_json(json_path)      # from our canonical JSON
    shot_list_io.write_xlsx(xlsx_path, data)      # produce the editable xlsx
    data = shot_list_io.read_xlsx(xlsx_path)      # read user edits back

Usage (CLI, generate xlsx from JSON):
    scripts/.venv/bin/python scripts/shot_list_io.py \\
        --from-json <session-dir>/shot-list/shot-list.json \\
        --to-xlsx <session-dir>/shot-list/shot-list.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from mutagen.mp3 import MP3  # type: ignore
    _HAS_MUTAGEN = True
except Exception:
    _HAS_MUTAGEN = False


REPO_ROOT = Path(__file__).resolve().parent.parent
CONTENT_ROOT = REPO_ROOT.parent / "spoolcast-content"


# ---- column spec -------------------------------------------------------

# scope: "chunk" columns use merged cells across the chunk's beats,
#        "beat" columns are per-row.
COLUMNS: list[tuple[str, str, str, int]] = [
    # (key, header, scope, width_chars)
    # Left: scannable chunk metadata.
    ("chunk_id",         "Chunk",            "chunk", 7),
    ("scene",            "Scene",            "chunk", 18),
    ("summary",          "Summary",          "chunk", 32),
    ("boundary_kind",    "Boundary",         "chunk", 14),
    ("weight",           "Weight",           "chunk", 8),
    ("act_title",        "Act Title",        "chunk", 12),
    ("continuity",       "Continuity",       "chunk", 20),
    ("reveal_direction", "Reveal Direction", "chunk", 14),
    # Beat data + its visual job, adjacent for side-by-side review.
    ("beat_id",          "Beat",             "beat",  7),
    ("narration",        "Narration",        "beat",  48),
    ("beat_description", "Beat Description", "chunk", 50),
    ("pause_after",      "Pause After",      "beat",  10),
    # Camera/timing (used when camera moves are planned; blank otherwise).
    ("camera_reason",    "Camera Reason",    "beat",  36),
    ("camera_target",    "Camera Target",    "beat",  14),
    ("camera_zoom",      "Camera Zoom",      "beat",  12),
    ("transition_s",     "Transition (s)",   "beat",  12),
    ("start_s",          "Start (s)",        "beat",  10),
    ("end_s",            "End (s)",          "beat",  10),
    # Right: reference + pacing data.
    ("context_justification", "Broll Context", "chunk", 42),
    ("audit_bridge",     "Audit: Bridge",    "beat",  30),
    ("audit_overweight", "Audit: Weight",    "beat",  24),
    ("full_prompt",      "Full Prompt",      "chunk", 60),
    ("image_source",     "Image Source",     "chunk", 14),
    ("image_path",       "Image Path",       "chunk", 36),
]

# Columns that should be hidden by default in the written xlsx. They're still
# present (for derivation/debugging) but collapsed so the sheet reads cleanly.
HIDDEN_COLUMN_KEYS = {"full_prompt"}

CHUNK_SCOPE_KEYS = {k for k, _, scope, _ in COLUMNS if scope == "chunk"}


# ---- palette -----------------------------------------------------------

CHUNK_ROW_COLORS = [
    "E3F2FD",  # pale blue
    "E8F5E9",  # pale green
    "FFF9E6",  # pale yellow
    "FCE4EC",  # pale pink
    "F3E5F5",  # pale purple
    "FFF3E0",  # pale orange
    "E0F2F1",  # pale teal
    "F5F5F5",  # pale gray
]


# ---- pause mapping -----------------------------------------------------

PAUSE_SECONDS = {
    "": 0.3,       # default rhythm — a brief natural beat
    "none": 0.0,
    "short": 0.3,
    "medium": 0.8,
    "long": 1.5,
}


# ---- data loading ------------------------------------------------------

def load_json(path: Path) -> dict[str, Any]:
    with open(path) as f:
        return json.load(f)


def _session_config(session_id: str) -> dict[str, Any]:
    cfg_path = CONTENT_ROOT / "sessions" / session_id / "session.json"
    if not cfg_path.exists():
        return {}
    with open(cfg_path) as f:
        return json.load(f)


def _audio_path(session_id: str, beat_id: str) -> Path:
    return CONTENT_ROOT / "sessions" / session_id / "source" / "audio" / f"{beat_id}.mp3"


def _audio_duration(session_id: str, beat_id: str) -> float | None:
    """Returns audio duration in seconds, or None if unavailable."""
    p = _audio_path(session_id, beat_id)
    if not p.exists() or not _HAS_MUTAGEN:
        return None
    try:
        return float(MP3(str(p)).info.length)
    except Exception:
        return None


def _compute_timeline(session_id: str, chunks: list[dict[str, Any]]) -> None:
    """Fill start_s/end_s on each beat based on audio duration + pause."""
    running = 0.0
    for chunk in chunks:
        for beat in chunk.get("beats", []):
            dur = _audio_duration(session_id, beat.get("id", ""))
            if dur is None:
                # fallback: estimate from narration length at ~14 cps
                text = beat.get("narration", "") or ""
                dur = max(1.0, len(text) / 14.0)
            pause = PAUSE_SECONDS.get(beat.get("pause_after", "") or "", 0.3)
            beat["_audio_duration"] = round(dur, 2)
            beat["start_s"] = round(running, 2)
            running += dur
            beat["end_s"] = round(running, 2)
            running += pause


# ---- prompt assembly ---------------------------------------------------

def _full_prompt(style_prompt: str, beat_description: str) -> str:
    style = (style_prompt or "").rstrip(".")
    beat = (beat_description or "").rstrip(".")
    parts = []
    if style:
        parts.append(style)
    if beat:
        parts.append(f"Scene: {beat}")
    return ". ".join(parts) + "." if parts else ""


# ---- xlsx writer -------------------------------------------------------

def write_xlsx(xlsx_path: Path, data: dict[str, Any]) -> None:
    session_id = data.get("session_id", "")
    canvas = data.get("canvas", {})
    chunks = data.get("chunks", [])

    cfg = _session_config(session_id)
    style_prompt = cfg.get("default_style_prompt") or cfg.get("style_reference") or ""

    # Compute timeline (fills start_s, end_s).
    _compute_timeline(session_id, chunks)

    # Load latest narration-audit report, if present, and map findings onto
    # beats/chunks so the xlsx surfaces them inline.
    audit_bridge_by_pair: dict[tuple[str, str], str] = {}
    audit_overweight_by_beat: dict[str, str] = {}
    audit_report_path = (
        CONTENT_ROOT / "sessions" / session_id / "working" / "narration-audit.json"
    )
    if audit_report_path.exists():
        try:
            with open(audit_report_path) as _f:
                _audit = json.load(_f)
            for _flag in _audit.get("bridge_flags", []):
                key = (_flag.get("beat_n_id", ""), _flag.get("beat_n1_id", ""))
                kind = _flag.get("connection_type") or _flag.get("verdict", "")
                proposed = _flag.get("proposed_bridge") or ""
                summary = f"[{kind}]"
                if proposed:
                    summary = f"{summary} {proposed}"
                audit_bridge_by_pair[key] = summary
            for _flag in _audit.get("overweight_flags", []):
                audit_overweight_by_beat[_flag.get("beat_id", "")] = _flag.get(
                    "proposed_fix", "overweight"
                )
        except (json.JSONDecodeError, OSError):
            pass

    wb = Workbook()
    ws = wb.active
    ws.title = "shot-list"

    # --- header area ---
    ws.cell(row=1, column=1, value=f"spoolcast shot list — {session_id}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

    meta_parts = [
        f"Session: {session_id}",
        f"Canvas: {canvas.get('aspect_ratio', '16:9')}",
        f"FPS: {canvas.get('fps', 30)}",
        f"Chunks: {len(chunks)}",
    ]
    ws.cell(row=2, column=1, value=" · ".join(meta_parts))
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    # row 3 left blank as spacer.

    # --- column headers (row 4) ---
    header_row = 4
    header_fill = PatternFill("solid", fgColor="D6D9DF")
    thin = Side(border_style="thin", color="CCCCCC")
    header_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx, (_key, label, _scope, _w) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = header_border

    # --- data rows ---
    data_start_row = header_row + 1
    current_row = data_start_row

    for chunk_idx, chunk in enumerate(chunks):
        beats = chunk.get("beats") or [{}]
        chunk_start = current_row
        chunk_end = current_row + len(beats) - 1
        color = CHUNK_ROW_COLORS[chunk_idx % len(CHUNK_ROW_COLORS)]
        fill = PatternFill("solid", fgColor=color)

        # Chunk-level values (written once; merged across beats).
        image_source = chunk.get("image_source", "generated")
        # For non-generated chunks, Full Prompt doesn't apply (no kie call).
        full_prompt_value = (
            _full_prompt(style_prompt, chunk.get("beat_description", ""))
            if image_source == "generated"
            else "—"
        )
        chunk_values = {
            "chunk_id": chunk.get("id", ""),
            "scene": f"{chunk.get('scene', '')} — {chunk.get('scene_title', '')}".strip(" —"),
            "summary": chunk.get("summary", ""),
            "boundary_kind": chunk.get("boundary_kind", ""),
            "weight": chunk.get("weight", ""),
            "act_title": chunk.get("act_title", ""),
            "context_justification": chunk.get("context_justification", ""),
            "continuity": chunk.get("continuity", "standalone"),
            "reveal_direction": chunk.get("reveal_direction", ""),
            "image_source": image_source,
            "image_path": chunk.get("image_path", ""),
            "beat_description": chunk.get("beat_description", ""),
            "full_prompt": full_prompt_value,
        }

        for beat_idx, beat in enumerate(beats):
            row = current_row + beat_idx
            for col_idx, (key, _label, scope, _w) in enumerate(COLUMNS, start=1):
                if scope == "chunk":
                    value = chunk_values.get(key, "") if beat_idx == 0 else None
                    v_align = "center"
                else:
                    # Beat-level: the "beat_id" column maps to beat["id"]
                    # (the JSON uses "id" per-beat).
                    if key == "beat_id":
                        value = beat.get("id", "")
                    elif key == "audit_bridge":
                        # Audit bridge flag lives on the pair (this beat, next beat).
                        this_id = beat.get("id", "")
                        next_id = beats[beat_idx + 1].get("id", "") if beat_idx + 1 < len(beats) else ""
                        # If the next beat is in a later chunk, look up that chunk's first beat.
                        if not next_id and chunk_idx + 1 < len(chunks):
                            next_chunk_beats = chunks[chunk_idx + 1].get("beats") or []
                            if next_chunk_beats:
                                next_id = next_chunk_beats[0].get("id", "")
                        value = audit_bridge_by_pair.get((this_id, next_id), "")
                    elif key == "audit_overweight":
                        value = audit_overweight_by_beat.get(beat.get("id", ""), "")
                    else:
                        value = beat.get(key, "")
                    # Normalize blank numeric fields to None so they don't
                    # show as "" stringly.
                    if key in ("transition_s", "start_s", "end_s") and value == "":
                        value = None
                    v_align = "top"
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.fill = fill
                cell.alignment = Alignment(vertical=v_align, wrap_text=True)
                cell.font = Font(size=11)
                # Red-tint bridge flag cells, orange-tint overweight cells so
                # flagged rows pop at a glance.
                if key == "audit_bridge" and value:
                    cell.fill = PatternFill("solid", fgColor="FFCDD2")
                    cell.font = Font(size=11, color="B71C1C")
                elif key == "audit_overweight" and value:
                    cell.fill = PatternFill("solid", fgColor="FFE0B2")
                    cell.font = Font(size=11, color="E65100")

        # Merge chunk-level columns across this chunk's beats.
        if chunk_end > chunk_start:
            for col_idx, (_key, _label, scope, _w) in enumerate(COLUMNS, start=1):
                if scope == "chunk":
                    ws.merge_cells(
                        start_row=chunk_start,
                        start_column=col_idx,
                        end_row=chunk_end,
                        end_column=col_idx,
                    )

        current_row = chunk_end + 1

    # --- column widths + hidden state ---
    for col_idx, (key, _label, _scope, width) in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
        if key in HIDDEN_COLUMN_KEYS:
            ws.column_dimensions[letter].hidden = True

    # --- row heights ---
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[header_row].height = 24
    # Data rows: fix a compact height so large cells (e.g. beat_description
    # or full_prompt) don't balloon the whole row. Excel/Numbers will still
    # wrap text inside the cell but won't auto-grow the row past this value.
    # If a specific chunk needs more room, the user can expand that row.
    for row in range(data_start_row, current_row):
        ws.row_dimensions[row].height = 42

    # --- freeze panes so header stays visible ---
    ws.freeze_panes = f"A{data_start_row}"

    # --- second sheet: asset checklist / QA pass ---
    _write_assets_sheet(wb, chunks, session_id)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


# ---- assets sheet (QA pass output) -------------------------------------

ASSET_COLUMNS: list[tuple[str, int]] = [
    ("Status",        8),
    ("Asset Path",    58),
    ("Type",          10),
    ("Purpose",       18),
    ("Used By",       22),
    ("Size",          10),
    ("Dims / Dur",    16),
    ("Concern",       46),
]


def _probe_duration(p: Path) -> float | None:
    import subprocess
    try:
        r = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(p)],
            capture_output=True, text=True, timeout=5,
        )
        return float(r.stdout.strip())
    except Exception:
        return None


def _image_dims(p: Path) -> tuple[int, int] | None:
    try:
        from PIL import Image
        with Image.open(p) as im:
            return im.size
    except Exception:
        return None


def _svg_has_geometry(p: Path) -> bool:
    import re
    try:
        text = p.read_text(errors="ignore")
    except Exception:
        return False
    return bool(re.search(r"<(path|g|polygon|rect|circle|ellipse)\b", text))


def _classify_asset(path_str: str) -> str:
    """Return asset type from extension."""
    ext = Path(path_str).suffix.lower()
    if ext in (".mp4", ".mov", ".webm", ".avi"): return "video"
    if ext == ".svg": return "svg"
    if ext in (".png", ".jpg", ".jpeg", ".webp", ".gif"): return "image"
    if ext in (".mp3", ".wav", ".m4a", ".ogg"): return "audio"
    return "other"


def _gather_external_assets(chunks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Collect unique external asset paths from chunks + overlays."""
    seen: dict[str, dict[str, Any]] = {}

    def add(path_str: str, purpose: str, used_by: str):
        if not path_str:
            return
        entry = seen.setdefault(path_str, {"path": path_str, "purposes": set(), "used_by": []})
        entry["purposes"].add(purpose)
        entry["used_by"].append(used_by)

    for chunk in chunks:
        src = chunk.get("image_source", "generated")
        img = chunk.get("image_path", "")
        # Only non-generated chunks contribute to the external-assets sheet.
        if src != "generated" and img:
            add(img, f"primary ({src})", chunk.get("id", ""))
        for ov in chunk.get("overlays", []) or []:
            ov_src = ov.get("source", "")
            if ov_src:
                add(ov_src, "overlay", chunk.get("id", ""))

    return list(seen.values())


def _write_assets_sheet(wb: "Workbook", chunks: list[dict[str, Any]], session_id: str) -> None:
    from openpyxl import Workbook  # noqa
    ws = wb.create_sheet("assets")

    # Title + meta
    total_cols = len(ASSET_COLUMNS)
    ws.cell(row=1, column=1, value=f"external assets checklist — {session_id}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=16)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)

    # Gather assets
    assets = _gather_external_assets(chunks)

    # Paths in shot-list.json are relative to the SESSION root (where session.json
    # and source/ live), NOT the shot-list/ subfolder. Cross-session refs use ../.
    session_root = CONTENT_ROOT / "sessions" / session_id

    # Meta line
    meta = f"Session: {session_id}  ·  External asset count: {len(assets)}  ·  This sheet is regenerated every time the shot list is written."
    ws.cell(row=2, column=1, value=meta)
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)

    # Headers at row 4
    header_fill = PatternFill("solid", fgColor="D6D9DF")
    thin = Side(border_style="thin", color="CCCCCC")
    header_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx, (label, _w) in enumerate(ASSET_COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=label)
        cell.font = Font(bold=True, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = header_border

    # Data rows
    OVERLAY_TARGET_PX = 200
    ok_fill   = PatternFill("solid", fgColor="E8F5E9")
    warn_fill = PatternFill("solid", fgColor="FFF3E0")
    fail_fill = PatternFill("solid", fgColor="FFEBEE")

    ok = warn = fail = 0

    for row_i, entry in enumerate(assets, start=5):
        rel = entry["path"]
        purpose = ", ".join(sorted(entry["purposes"]))
        used_by = ", ".join(dict.fromkeys(entry["used_by"]))  # dedupe preserving order
        atype = _classify_asset(rel)

        p = (session_root / rel).resolve() if not rel.startswith("/") else Path(rel)
        if not p.exists():
            status = "❌"
            size_str = "MISSING"
            dims_str = ""
            concern = "file not on disk"
        else:
            sz = p.stat().st_size
            size_str = f"{sz // 1024} KB" if sz >= 1024 else f"{sz} B"
            concerns: list[str] = []
            dims_str = ""

            if atype == "image":
                dims = _image_dims(p)
                if dims is None:
                    concerns.append("unreadable image")
                else:
                    w, h = dims
                    dims_str = f"{w}x{h}"
                    if "overlay" in purpose:
                        if max(w, h) < OVERLAY_TARGET_PX:
                            concerns.append(f"{max(w,h)}px < {OVERLAY_TARGET_PX}px overlay target")
            elif atype == "svg":
                dims_str = "svg"
                if not _svg_has_geometry(p):
                    concerns.append("no geometry found")
            elif atype == "video":
                dur = _probe_duration(p)
                dims_str = f"{dur:.1f}s" if dur else "?"
                if sz < 50_000:
                    concerns.append(f"small file ({sz}B)")
            elif atype == "audio":
                dur = _probe_duration(p)
                dims_str = f"{dur:.1f}s" if dur else "?"
                if dur is not None and dur < 0.5:
                    concerns.append("audio too short")

            if concerns:
                status = "⚠️"
            else:
                status = "✅"
            concern = "; ".join(concerns) if concerns else "ok"

        if status == "✅":   ok += 1;   row_fill = ok_fill
        elif status == "⚠️": warn += 1; row_fill = warn_fill
        else:                fail += 1; row_fill = fail_fill

        values = [status, rel, atype, purpose, used_by, size_str, dims_str, concern]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = Font(size=11)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Summary row at the bottom
    summary_row = 5 + len(assets) + 1
    total = ok + warn + fail
    summary = f"SUMMARY: {ok} ✅   {warn} ⚠️   {fail} ❌   of {total} total"
    ws.cell(row=summary_row, column=1, value=summary)
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=total_cols)

    # Column widths
    for col_idx, (_label, width) in enumerate(ASSET_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row heights
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 22
    for r in range(5, 5 + len(assets)):
        ws.row_dimensions[r].height = 28

    # Freeze panes
    ws.freeze_panes = "A5"


# ---- xlsx reader -------------------------------------------------------

def read_xlsx(xlsx_path: Path) -> dict[str, Any]:
    """Read the xlsx back into the normalized data dict.

    Merged chunk-level cells: the top-left of the merge holds the value;
    other cells in the merge are None. We forward-fill those by chunk.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Find the header row by looking for the "Chunk" label.
    header_row = None
    for r in range(1, 10):
        if ws.cell(row=r, column=1).value == "Chunk":
            header_row = r
            break
    if header_row is None:
        raise ValueError(f"could not locate header row in {xlsx_path}")

    key_by_col = {}
    for col_idx, (key, label, _scope, _w) in enumerate(COLUMNS, start=1):
        key_by_col[col_idx] = key

    chunks: dict[str, dict[str, Any]] = {}
    chunk_order: list[str] = []

    # Build a value lookup that resolves merged cells to the top-left value.
    def _get(row: int, col: int) -> Any:
        cell = ws.cell(row=row, column=col)
        # if the cell is part of a merged range, return the top-left value
        for merged in ws.merged_cells.ranges:
            if cell.coordinate in merged:
                return ws.cell(row=merged.min_row, column=merged.min_col).value
        return cell.value

    r = header_row + 1
    while True:
        chunk_id = _get(r, 1)
        if not chunk_id:
            break
        beat_id = _get(r, 4)  # "Beat" column

        if chunk_id not in chunks:
            # New chunk: capture chunk-level fields by looking up by key name.
            scene_raw = _get_by_key(r, "scene", key_by_col, _get) or ""
            if " — " in scene_raw:
                scene, scene_title = scene_raw.split(" — ", 1)
            else:
                scene, scene_title = scene_raw, ""
            chunks[chunk_id] = {
                "id": chunk_id,
                "scene": scene,
                "scene_title": scene_title,
                "summary": _get_by_key(r, "summary", key_by_col, _get) or "",
                "continuity": _get_by_key(r, "continuity", key_by_col, _get) or "standalone",
                "reveal_direction": _get_by_key(r, "reveal_direction", key_by_col, _get) or "",
                "image_source": _get_by_key(r, "image_source", key_by_col, _get) or "generated",
                "image_path": _get_by_key(r, "image_path", key_by_col, _get) or "",
                "beat_description": _get_by_key(r, "beat_description", key_by_col, _get) or "",
                "beats": [],
            }
            chunk_order.append(chunk_id)

        beat: dict[str, Any] = {"id": beat_id or ""}
        for col_idx in range(1, len(COLUMNS) + 1):
            key = key_by_col[col_idx]
            if key in CHUNK_SCOPE_KEYS or key == "beat_id":
                continue
            value = ws.cell(row=r, column=col_idx).value
            if value is None:
                continue
            beat[key] = value
        chunks[chunk_id]["beats"].append(beat)
        r += 1

    return {
        "session_id": _read_session_id_from_meta(ws),
        "canvas": _read_canvas_from_meta(ws),
        "chunks": [chunks[cid] for cid in chunk_order],
    }


def _get_by_key(row: int, key: str, key_by_col: dict[int, str], getter: Any) -> Any:
    for col_idx, k in key_by_col.items():
        if k == key:
            return getter(row, col_idx)
    return None


def _read_session_id_from_meta(ws: Any) -> str:
    meta = ws.cell(row=2, column=1).value or ""
    for part in str(meta).split("·"):
        part = part.strip()
        if part.lower().startswith("session:"):
            return part.split(":", 1)[1].strip()
    return ""


def _read_canvas_from_meta(ws: Any) -> dict[str, Any]:
    meta = ws.cell(row=2, column=1).value or ""
    canvas: dict[str, Any] = {}
    for part in str(meta).split("·"):
        part = part.strip()
        if part.lower().startswith("canvas:"):
            canvas["aspect_ratio"] = part.split(":", 1)[1].strip()
        elif part.lower().startswith("fps:"):
            try:
                canvas["fps"] = int(part.split(":", 1)[1].strip())
            except Exception:
                pass
    return canvas


# ---- CLI ---------------------------------------------------------------

def _cli() -> None:
    parser = argparse.ArgumentParser(
        description="Convert between shot-list JSON and xlsx"
    )
    sub = parser.add_subparsers(dest="cmd", required=True)

    p1 = sub.add_parser("to-xlsx", help="Generate xlsx from JSON")
    p1.add_argument("--from-json", required=True)
    p1.add_argument("--to-xlsx", required=True)

    p2 = sub.add_parser("from-xlsx", help="Read xlsx and dump normalized JSON to stdout")
    p2.add_argument("--xlsx", required=True)

    args = parser.parse_args()

    if args.cmd == "to-xlsx":
        data = load_json(Path(args.from_json))
        write_xlsx(Path(args.to_xlsx), data)
        print(f"[shot-list] wrote {args.to_xlsx}")
    elif args.cmd == "from-xlsx":
        data = read_xlsx(Path(args.xlsx))
        print(json.dumps(data, indent=2))


if __name__ == "__main__":
    _cli()
